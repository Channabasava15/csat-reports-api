# import numpy as np
# import pandas as pd
# from flask import Flask, request, jsonify
# from flask_cors import CORS
# import asyncio
# import hashlib
# import hmac
# import base64
# import time
# import aiohttp
# from aiohttp import ClientTimeout
# from urllib.parse import urlparse, parse_qs, urlencode
# import pandas as pd
# from datetime import datetime
# import os
# import re
# from openpyxl import load_workbook
# from io import BytesIO
# import smtplib
# from email.message import EmailMessage
# import ssl
# import requests.utils
# import threading
# import uuid
# import signal
# import sys

# app = Flask(__name__)
# CORS(app)

# # Store job status
# jobs = {}

# # ===============================
# # CONFIG
# # ===============================
# config = {
#     'baseUrl': 'https://3373505-sb1.restlets.api.netsuite.com/app/site/hosting/restlet.nl',
#     'realm': '3373505_SB1',
#     'consumerKey': 'c6d5e52b1d6b4fa3843170179f35f9eaa725893f03df5191c7b364486bd78253',
#     'consumerSecret': 'b3d567c88d518d3db217c6a3c5281684b418a96d471aacd9891d28b4e71394db',
#     'token': '82e2861cb21ac4092726b1fe007c736a4616fa72a35e5df4af92e0773541bd32',
#     'tokenSecret': '6d36cd7ba2d246100a82c892f543b7eabdea61bcf3a0be691849e42538bcb0e4'
# }

# # Email configuration
# EMAIL_CONFIG = {
#     'sender': 'channabasava@ayurvaid.com',
#     'password': 'tnpyncniqefpcaqn',
#     'smtp_server': 'smtp.mail.yahoo.com',
#     'smtp_port': 465,
#     'receivers': ['channabasavah50@gmail.com']
# }


# # Search configurations
# treatment_searches = [
#     {'id': 1763, 'name': 'Data-SO'},
#     {'id': 1710, 'name': 'Data-TSD+1'},
#     {'id': 1712, 'name': 'Data-TED-1'}
# ]

# consultation_searches = [
#     {'id': 1719, 'name': 'C+8 Data'},
#     {'id': 1764, 'name': 'Data-PP1'}
# ]

# CENTRE_MAPPING_SEARCH_ID = 1765

# # ===============================
# # HELPER FUNCTIONS
# # ===============================
# def format_month_year(date_str):
#     """Convert DD/MM/YYYY to MMM-YY format (e.g., 01/01/2025 -> Jan-25)"""
#     try:
#         date_obj = datetime.strptime(date_str, '%d/%m/%Y')
#         return date_obj.strftime('%b-%y')
#     except:
#         return date_str

# # ===============================
# # OAUTH FUNCTIONS
# # ===============================
# def generate_oauth_signature(url, method, consumer_key, consumer_secret, token_key, token_secret, realm):
#     parsed_url = urlparse(url)
#     query_params = parse_qs(parsed_url.query)

#     oauth_params = {
#         'oauth_consumer_key': consumer_key,
#         'oauth_nonce': base64.b64encode(os.urandom(32)).decode()[:32],
#         'oauth_signature_method': 'HMAC-SHA256',
#         'oauth_timestamp': str(int(time.time())),
#         'oauth_token': token_key,
#         'oauth_version': '1.0'
#     }

#     all_params = {k: v[0] for k, v in query_params.items()}
#     all_params.update(oauth_params)

#     sorted_params = sorted(all_params.items())
#     param_string = '&'.join([f"{k}={requests.utils.quote(str(v), safe='')}" for k, v in sorted_params])

#     base_string = '&'.join([
#         method.upper(),
#         requests.utils.quote(parsed_url.scheme + '://' + parsed_url.netloc + parsed_url.path, safe=''),
#         requests.utils.quote(param_string, safe='')
#     ])

#     signing_key = f"{requests.utils.quote(consumer_secret, safe='')}&{requests.utils.quote(token_secret, safe='')}"

#     signature = base64.b64encode(
#         hmac.new(signing_key.encode(), base_string.encode(), hashlib.sha256).digest()
#     ).decode()

#     oauth_params['oauth_signature'] = signature

#     auth_header = 'OAuth realm="' + realm + '", ' + ', '.join([
#         f'{k}="{requests.utils.quote(str(v), safe="")}"' for k, v in oauth_params.items()
#     ])

#     return {'Authorization': auth_header}

# def get_auth_header(url):
#     return generate_oauth_signature(
#         url, 'GET',
#         config['consumerKey'],
#         config['consumerSecret'],
#         config['token'],
#         config['tokenSecret'],
#         config['realm']
#     )

# # ===============================
# # FETCH DATA FUNCTIONS
# # ===============================
# async def fetch_with_retry(session, url, retries=3):
#     for attempt in range(retries):
#         try:
#             headers = get_auth_header(url)
#             headers.update({'Content-Type': 'application/json'})

#             async with session.get(url, headers=headers) as res:
#                 if res.status == 200:
#                     return await res.json()
#                 else:
#                     error_text = await res.text()
#                     print(f"   HTTP {res.status}: {error_text[:100]}")
#         except Exception as e:
#             print(f"   Attempt {attempt + 1} failed: {str(e)[:50]}")
#             await asyncio.sleep(2)
#     return None

# async def fetch_search(session, search_id, start_date, end_date):
#     print(f"\n🚀 Fetching {search_id} (Date range: {start_date} to {end_date})")
#     all_data = []
#     page = 1

#     while True:
#         params = {
#             'script': 559,
#             'deploy': 1,
#             'searchId': search_id,
#             'pageSize': 100,
#             'pageNumber': page,
#             'startDate': start_date,
#             'endDate': end_date
#         }
        
#         url = f"{config['baseUrl']}?{urlencode(params)}"
#         res = await fetch_with_retry(session, url)

#         if not res:
#             print(f"🛑 No response for page {page}")
#             break
            
#         if not res.get('results'):
#             print(f"🛑 No results at page {page}")
#             break

#         results = res['results']
#         all_data.extend(results)
#         print(f"📄 Page {page} → {len(results)} rows")

#         if len(results) < 100:
#             break
#         page += 1

#     print(f"✅ Total for {search_id}: {len(all_data)} rows")
#     return all_data

# async def fetch_centre_mapping(session):
#     """Fetch centre mapping and return both mapping dict and list of centres"""
#     print(f"\n🏥 Fetching centre recipient mapping")
    
#     params = {
#         'script': 559,
#         'deploy': 1,
#         'searchId': CENTRE_MAPPING_SEARCH_ID,
#         'pageSize': 100,
#         'pageNumber': 1
#     }
    
#     url = f"{config['baseUrl']}?{urlencode(params)}"
#     res = await fetch_with_retry(session, url)

#     if not res or not res.get('results'):
#         print("❌ No centre mapping data found!")
#         return {}, []
    
#     centre_mapping = {}
#     centre_list = []
    
#     for item in res['results']:
#         centre_name = item.get('AyurVAID Centre', '').strip()
#         if not centre_name:
#             continue
            
#         recipients = []
#         if item.get('Recipient 1'):
#             recipients.append(item['Recipient 1'].strip())
#         if item.get('Recipient 2'):
#             recipients.append(item['Recipient 2'].strip())
        
#         if recipients:
#             centre_mapping[centre_name] = recipients
#             centre_list.append(centre_name)
#             print(f"   📍 {centre_name}: {', '.join(recipients)}")
    
#     centre_list.sort()
#     print(f"✅ Loaded centre mapping for {len(centre_mapping)} centres")
    
#     return centre_mapping, centre_list

# async def fetch_filtered_data(report_type, start_date, end_date):
#     timeout = ClientTimeout(total=300)
    
#     async with aiohttp.ClientSession(timeout=timeout) as session:
#         centre_mapping, centre_list = await fetch_centre_mapping(session)
        
#         treatment_data = {}
#         consultation_data = {}
        
#         if report_type in ['treatment', 'both']:
#             print("\n📋 FETCHING TREATMENT REPORT DATA")
#             for s in treatment_searches:
#                 treatment_data[s['name']] = await fetch_search(session, s['id'], start_date, end_date)
        
#         if report_type in ['consultation', 'both']:
#             print("\n📋 FETCHING CONSULTATION REPORT DATA")
#             for s in consultation_searches:
#                 consultation_data[s['name']] = await fetch_search(session, s['id'], start_date, end_date)
        
#         return centre_mapping, centre_list, treatment_data, consultation_data

# # ===============================
# # REPORT CREATION FUNCTIONS
# # ===============================
# def adjust_formula(formula, row):
#     return re.sub(r'(\$?[A-Z]+)(\$?)2', lambda m: f"{m.group(1)}{m.group(2)}{row}", formula)

# def create_report_in_memory(all_data, template_path, sheet_names, start_date=None, target_sheet_for_date=None):
#     """Create report from data with date in specific sheet's F1 cell"""
#     if not all_data or all(all_data.get(name) == [] for name in sheet_names):
#         print(f"⚠️ No data to create report")
#         return None
    
#     wb = load_workbook(template_path)
    
#     # Set the formatted date in the target sheet's F1 cell
#     if start_date and target_sheet_for_date:
#         formatted_date = format_month_year(start_date)
#         if target_sheet_for_date in wb.sheetnames:
#             ws = wb[target_sheet_for_date]
#             ws.cell(row=1, column=6, value=formatted_date)
#             print(f"   📅 Set date in {target_sheet_for_date}!F1 to: {formatted_date}")
#         else:
#             print(f"   ⚠️ Target sheet '{target_sheet_for_date}' not found in template")
    
#     for name in sheet_names:
#         if name not in wb.sheetnames:
#             print(f"⚠️ Sheet '{name}' not found in template")
#             continue
        
#         ws = wb[name]
        
#         formulas = {}
#         for col in range(1, ws.max_column + 1):
#             val = ws.cell(row=2, column=col).value
#             if isinstance(val, str) and val.startswith("="):
#                 formulas[col] = val
        
#         ws.delete_rows(3, ws.max_row)
        
#         data = all_data.get(name, [])
#         if not data:
#             print(f"⚠️ No data for {name}")
#             continue
        
#         df = pd.DataFrame(data).fillna('')
#         print(f"📝 Writing {name}: {len(df)} rows")
        
#         for r_idx, row in enumerate(df.values, start=2):
#             for c_idx, val in enumerate(row, start=1):
#                 ws.cell(row=r_idx, column=c_idx, value=val)
        
#         for col, formula in formulas.items():
#             for r in range(2, len(df) + 2):
#                 ws.cell(row=r, column=col, value=adjust_formula(formula, r))
    
#     wb.calculation.fullCalcOnLoad = True
    
#     output = BytesIO()
#     wb.save(output)
#     output.seek(0)
#     return output

# def filter_data_by_centre(all_data, centre_name, centre_column_name="AyurVAID Centre"):
#     """Filter data for a specific centre"""
#     filtered_data = {}
    
#     for sheet_name, data_list in all_data.items():
#         if not data_list:
#             filtered_data[sheet_name] = []
#             continue
        
#         df = pd.DataFrame(data_list)
        
#         if centre_column_name in df.columns:
#             filtered_df = df[df[centre_column_name] == centre_name]
#             filtered_data[sheet_name] = filtered_df.to_dict('records')
#             print(f"   📊 {sheet_name}: {len(filtered_df)} records for {centre_name}")
#         else:
#             print(f"   ⚠️ Column '{centre_column_name}' not found in {sheet_name}")
#             filtered_data[sheet_name] = []
    
#     return filtered_data

# def send_email_with_attachment(recipients, subject, body, attachments, job_id=None):
#     """Send email with attachments"""
#     if not recipients:
#         print("   ⚠️ No recipients specified")
#         return False
    
#     msg = EmailMessage()
#     msg["Subject"] = subject
#     msg["From"] = EMAIL_CONFIG['sender']
#     msg["To"] = ", ".join(recipients)
    
#     # Add BCC to default receivers for tracking
#     if EMAIL_CONFIG['receivers']:
#         msg["Bcc"] = ", ".join(EMAIL_CONFIG['receivers'])
    
#     msg.set_content(body)
    
#     # Attach files
#     for filename, data in attachments:
#         data.seek(0)
#         msg.add_attachment(
#             data.read(),
#             maintype="application",
#             subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#             filename=filename
#         )
    
#     # Send email with retry logic
#     max_retries = 3
#     for attempt in range(max_retries):
#         try:
#             print(f"   📤 Attempting to send (attempt {attempt + 1}/{max_retries})...")
            
#             context = ssl.create_default_context()
            
#             with smtplib.SMTP_SSL(EMAIL_CONFIG['smtp_server'], EMAIL_CONFIG['smtp_port'], context=context) as smtp:
#                 smtp.set_debuglevel(0)
#                 smtp.login(EMAIL_CONFIG['sender'], EMAIL_CONFIG['password'])
#                 smtp.send_message(msg)
#             # context = ssl.create_default_context()
            
#             # # Change to standard SMTP and explicitly use port 587
#             # with smtplib.SMTP(EMAIL_CONFIG['smtp_server'], 587) as smtp:
#             #     smtp.set_debuglevel(0)
#             #     smtp.starttls(context=context) # Upgrade the connection to secure
#             #     smtp.login(EMAIL_CONFIG['sender'], EMAIL_CONFIG['password'])
#             #     smtp.send_message(msg)
            
#             print(f"   ✅ Email sent successfully to {', '.join(recipients)}")
#             return True
            
#         except Exception as e:
#             print(f"   ❌ Attempt {attempt + 1} failed: {str(e)}")
#             if attempt < max_retries - 1:
#                 wait_time = 5 * (attempt + 1)
#                 print(f"   Waiting {wait_time} seconds before retry...")
#                 time.sleep(wait_time)
    
#     print(f"   ❌ Failed to send email after {max_retries} attempts")
#     return False




# def create_centre_wise_reports(all_data, template_path, sheet_names, report_prefix, timestamp, start_date, target_sheet_for_date=None):
#     """Create centre-wise reports with date in specific sheet's F1"""
#     centre_data = {}
    
#     for sheet_name, data_list in all_data.items():
#         if not data_list:
#             continue
        
#         df = pd.DataFrame(data_list)
        
#         if "AyurVAID Centre" not in df.columns:
#             print(f"⚠️ Column 'AyurVAID Centre' not found in {sheet_name}")
#             continue
        
#         for centre, group_df in df.groupby("AyurVAID Centre"):
#             if pd.isna(centre) or centre == "":
#                 centre = "Unknown_Centre"
            
#             if centre not in centre_data:
#                 centre_data[centre] = {}
            
#             centre_data[centre][sheet_name] = group_df
    
#     if not centre_data:
#         print("⚠️ No centre-wise data found!")
#         return []
    
#     centre_files = []
#     formatted_date = format_month_year(start_date)
    
#     for centre, sheets_data in centre_data.items():
#         centre_clean = str(centre).replace("/", "_").replace("\\", "_").replace(" ", "_")
#         filename = f"{report_prefix}_{centre_clean}_{timestamp}.xlsx"
        
#         wb = load_workbook(template_path)
        
#         # Set the formatted date in the target sheet's F1 cell
#         if target_sheet_for_date and target_sheet_for_date in wb.sheetnames:
#             ws = wb[target_sheet_for_date]
#             ws.cell(row=1, column=6, value=formatted_date)
#             print(f"   📅 Set date in {target_sheet_for_date}!F1 to: {formatted_date} for {centre}")
        
#         for sheet_name in sheet_names:
#             if sheet_name not in wb.sheetnames:
#                 continue
            
#             ws = wb[sheet_name]
            
#             formulas = {}
#             for col in range(1, ws.max_column + 1):
#                 val = ws.cell(row=2, column=col).value
#                 if isinstance(val, str) and val.startswith("="):
#                     formulas[col] = val
            
#             ws.delete_rows(3, ws.max_row)
            
#             data = sheets_data.get(sheet_name, pd.DataFrame())
#             if data.empty:
#                 continue
            
#             df = data.reset_index(drop=True).fillna('')
#             print(f"   📝 {centre} - {sheet_name}: {len(df)} rows")
            
#             for r_idx, row in enumerate(df.values, start=2):
#                 for c_idx, val in enumerate(row, start=1):
#                     ws.cell(row=r_idx, column=c_idx, value=val)
            
#             for col, formula in formulas.items():
#                 for r in range(2, len(df) + 2):
#                     ws.cell(row=r, column=col, value=adjust_formula(formula, r))
        
#         wb.calculation.fullCalcOnLoad = True
        
#         output = BytesIO()
#         wb.save(output)
#         output.seek(0)
        
#         centre_files.append({
#             'centre_name': centre,
#             'filename': filename,
#             'data': output
#         })
    
#     print(f"✅ Created {len(centre_files)} centre-wise reports with date {formatted_date}")
#     return centre_files

# # ===============================
# # BACKGROUND JOB FUNCTION
# # ===============================
# def run_report_generation(report_type, start_date, end_date, selected_centre, job_id):
#     try:
#         jobs[job_id]['status'] = 'fetching_data'
#         jobs[job_id]['message'] = 'Fetching data from NetSuite...'
#         jobs[job_id]['progress'] = 20
        
#         # Fetch data
#         centre_mapping, centre_list, treatment_data, consultation_data = asyncio.run(
#             fetch_filtered_data(report_type, start_date, end_date)
#         )
        
#         if not centre_mapping:
#             print("\n⚠️ No centre mapping found!")
#             jobs[job_id]['status'] = 'failed'
#             jobs[job_id]['message'] = 'No centre mapping found'
#             return
        
#         jobs[job_id]['progress'] = 40
#         jobs[job_id]['message'] = 'Creating reports...'
        
#         timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
#         date_range_str = f"{start_date} to {end_date}"
        
#         # Handle specific centre or ALL
#         if selected_centre == 'ALL':
#             # Generate centre-wise reports for all centres
#             treatment_centre_files = []
#             consultation_centre_files = []
            
#             # Create centre-wise treatment reports
#             if report_type in ['treatment', 'both'] and treatment_data:
#                 treatment_centre_files = create_centre_wise_reports(
#                     treatment_data, 
#                     "CSAT_Treatment_template.xlsx", 
#                     ["Data-SO", "Data-TSD+1", "Data-TED-1"], 
#                     "Treatment_Report", 
#                     timestamp, 
#                     start_date,
#                     "CSAT-AVH"
#                 )
            
#             # Create centre-wise consultation reports
#             if report_type in ['consultation', 'both'] and consultation_data:
#                 consultation_centre_files = create_centre_wise_reports(
#                     consultation_data, 
#                     "CSAT_Consultation_template.xlsx",
#                     ["C+8 Data", "Data-PP1"],
#                     "Consultation_Report", 
#                     timestamp, 
#                     start_date,
#                     "C+8 Dashboard"
#                 )
            
#             # Send emails to all centres
#             jobs[job_id]['progress'] = 60
#             jobs[job_id]['message'] = 'Sending centre-wise emails...'
            
#             successful_sends = 0
#             for centre_name, recipients in centre_mapping.items():
#                 attachments = []
                
#                 # Find treatment report for this centre
#                 treatment_file = next((f for f in treatment_centre_files if f['centre_name'] == centre_name), None)
#                 if treatment_file:
#                     attachments.append((treatment_file['filename'], treatment_file['data']))
                
#                 # Find consultation report for this centre
#                 consultation_file = next((f for f in consultation_centre_files if f['centre_name'] == centre_name), None)
#                 if consultation_file:
#                     attachments.append((consultation_file['filename'], consultation_file['data']))
                
#                 if attachments:
#                     subject = f"CSAT Reports - {centre_name} ({date_range_str})"
#                     body = f"""Dear Team,

# Please find attached CSAT reports for {centre_name} centre for the date range: {date_range_str}

# Regards,
# Automation Bot"""
                    
#                     if send_email_with_attachment(recipients, subject, body, attachments, job_id):
#                         successful_sends += 1
#                     time.sleep(2)  # Delay between emails
            
#             jobs[job_id]['message'] = f'Emails sent to {successful_sends} centres'
            
#         else:
#             # Filter data for specific centre
#             jobs[job_id]['message'] = f'Filtering data for {selected_centre}...'
            
#             # Filter treatment data
#             filtered_treatment_data = None
#             if report_type in ['treatment', 'both'] and treatment_data:
#                 filtered_treatment_data = filter_data_by_centre(treatment_data, selected_centre)
            
#             # Filter consultation data
#             filtered_consultation_data = None
#             if report_type in ['consultation', 'both'] and consultation_data:
#                 filtered_consultation_data = filter_data_by_centre(consultation_data, selected_centre)
            
#             # Create reports for specific centre
#             attachments = []
            
#             if filtered_treatment_data:
#                 treatment_filename = f"Treatment_Report_{selected_centre}_{start_date.replace('/', '-')}_to_{end_date.replace('/', '-')}_{timestamp}.xlsx"
#                 treatment_data_bytes = create_report_in_memory(
#                     filtered_treatment_data, 
#                     "CSAT_Treatment_template.xlsx",
#                     ["Data-SO", "Data-TSD+1", "Data-TED-1"],
#                     start_date,
#                     "CSAT-AVH"
#                 )
#                 if treatment_data_bytes:
#                     attachments.append((treatment_filename, treatment_data_bytes))
#                     print(f"✅ Created treatment report for {selected_centre}")
            
#             if filtered_consultation_data:
#                 consultation_filename = f"Consultation_Report_{selected_centre}_{start_date.replace('/', '-')}_to_{end_date.replace('/', '-')}_{timestamp}.xlsx"
#                 consultation_data_bytes = create_report_in_memory(
#                     filtered_consultation_data, 
#                     "CSAT_Consultation_template.xlsx",
#                     ["C+8 Data", "Data-PP1"],
#                     start_date,
#                     "C+8 Dashboard"
#                 )
#                 if consultation_data_bytes:
#                     attachments.append((consultation_filename, consultation_data_bytes))
#                     print(f"✅ Created consultation report for {selected_centre}")
            
#             # Send email to specific centre
#             if selected_centre in centre_mapping:
#                 recipients = centre_mapping[selected_centre]
#                 subject = f"CSAT Reports - {selected_centre} ({date_range_str})"
#                 body = f"""Dear Team,

# Please find attached CSAT reports for {selected_centre} centre for the date range: {date_range_str}

# Regards,
# Automation Bot"""
                
#                 jobs[job_id]['progress'] = 60
#                 jobs[job_id]['message'] = f'Sending email to {selected_centre}...'
                
#                 if send_email_with_attachment(recipients, subject, body, attachments, job_id):
#                     jobs[job_id]['message'] = f'Email sent to {selected_centre}'
#                 else:
#                     jobs[job_id]['message'] = f'Failed to send email to {selected_centre}'
#             else:
#                 jobs[job_id]['message'] = f'Centre {selected_centre} not found in mapping'
#                 print(f"❌ Centre '{selected_centre}' not found in centre mapping")
        
#         # Send master email to admins
#         jobs[job_id]['progress'] = 80
#         jobs[job_id]['message'] = 'Sending master email to admins...'
        
#         # Create master reports
#         master_attachments = []
        
#         if report_type in ['treatment', 'both'] and treatment_data:
#             master_treatment_filename = f"Master_Treatment_Report_{start_date.replace('/', '-')}_to_{end_date.replace('/', '-')}_{timestamp}.xlsx"
#             master_treatment_data = create_report_in_memory(
#                 treatment_data, 
#                 "CSAT_Treatment_template.xlsx",
#                 ["Data-SO", "Data-TSD+1", "Data-TED-1"],
#                 start_date,
#                 "CSAT-AVH"
#             )
#             if master_treatment_data:
#                 master_attachments.append((master_treatment_filename, master_treatment_data))
        
#         if report_type in ['consultation', 'both'] and consultation_data:
#             master_consultation_filename = f"Master_Consultation_Report_{start_date.replace('/', '-')}_to_{end_date.replace('/', '-')}_{timestamp}.xlsx"
#             master_consultation_data = create_report_in_memory(
#                 consultation_data, 
#                 "CSAT_Consultation_template.xlsx",
#                 ["C+8 Data", "Data-PP1"],
#                 start_date,
#                 "C+8 Dashboard"
#             )
#             if master_consultation_data:
#                 master_attachments.append((master_consultation_filename, master_consultation_data))
        
#         if master_attachments:
#             master_subject = f"CSAT Reports - MASTER ({date_range_str})"
#             master_body = f"""Hi,

# Please find attached master CSAT reports for date range: {date_range_str}

# Regards,
# Automation Bot"""
            
#             send_email_with_attachment(EMAIL_CONFIG['receivers'], master_subject, master_body, master_attachments, job_id)
        
#         jobs[job_id]['status'] = 'completed'
#         jobs[job_id]['message'] = 'Report generation completed successfully!'
#         jobs[job_id]['progress'] = 100
        
#         print("\n" + "="*60)
#         print("✅ JOB COMPLETED SUCCESSFULLY")
#         print("="*60)
        
#     except Exception as e:
#         jobs[job_id]['status'] = 'failed'
#         jobs[job_id]['message'] = f'Error: {str(e)}'
#         jobs[job_id]['progress'] = 0
#         print(f"\n❌ Error in job {job_id}: {str(e)}")
#         import traceback
#         traceback.print_exc()

# # ===============================
# # FLASK ROUTES (API ONLY)
# # ===============================

# @app.route('/health', methods=['GET'])
# def health_check():
#     """Health check endpoint for Render"""
#     return jsonify({'status': 'healthy', 'timestamp': datetime.now().isoformat()})

# @app.route('/api/get_centres', methods=['GET'])
# def get_centres():
#     """Fetch list of centres for dropdown"""
#     async def fetch_centres():
#         timeout = ClientTimeout(total=30)
#         async with aiohttp.ClientSession(timeout=timeout) as session:
#             _, centre_list = await fetch_centre_mapping(session)
#             return centre_list
    
#     try:
#         centre_list = asyncio.run(fetch_centres())
#         return jsonify({'centres': centre_list, 'count': len(centre_list)})
#     except Exception as e:
#         return jsonify({'error': str(e), 'centres': []}), 500

# @app.route('/api/generate_report', methods=['POST'])
# def generate_report():
#     """Generate and send CSAT reports"""
#     data = request.json
    
#     if not data:
#         return jsonify({'error': 'No JSON data provided'}), 400
    
#     report_type = data.get('report_type')
#     start_date = data.get('start_date')
#     end_date = data.get('end_date')
#     selected_centre = data.get('selected_centre', 'ALL')
    
#     # Validate inputs
#     if not report_type or not start_date or not end_date:
#         return jsonify({'error': 'Missing required parameters: report_type, start_date, end_date'}), 400
    
#     if report_type not in ['treatment', 'consultation', 'both']:
#         return jsonify({'error': 'Invalid report_type. Must be: treatment, consultation, or both'}), 400
    
#     # Validate date format
#     try:
#         datetime.strptime(start_date, '%d/%m/%Y')
#         datetime.strptime(end_date, '%d/%m/%Y')
#     except ValueError:
#         return jsonify({'error': 'Invalid date format. Use DD/MM/YYYY'}), 400
    
#     # Create job
#     job_id = str(uuid.uuid4())
    
#     jobs[job_id] = {
#         'status': 'pending',
#         'message': 'Job started',
#         'progress': 0,
#         'report_type': report_type,
#         'start_date': start_date,
#         'end_date': end_date,
#         'selected_centre': selected_centre,
#         'created_at': datetime.now().isoformat()
#     }
    
#     # Start background thread
#     thread = threading.Thread(
#         target=run_report_generation,
#         args=(report_type, start_date, end_date, selected_centre, job_id)
#     )
#     thread.daemon = True
#     thread.start()
    
#     return jsonify({
#         'job_id': job_id, 
#         'status': 'started',
#         'message': 'Report generation started. Use /api/job_status/<job_id> to check progress'
#     })

# @app.route('/api/job_status/<job_id>', methods=['GET'])
# def job_status(job_id):
#     """Get status of a specific job"""
#     if job_id not in jobs:
#         return jsonify({'error': 'Job not found'}), 404
    
#     return jsonify(jobs[job_id])

# @app.route('/api/jobs', methods=['GET'])
# def list_jobs():
#     """List all jobs (optional, useful for monitoring)"""
#     return jsonify({
#         'total_jobs': len(jobs),
#         'jobs': {job_id: {
#             'status': job['status'],
#             'report_type': job.get('report_type'),
#             'start_date': job.get('start_date'),
#             'end_date': job.get('end_date'),
#             'selected_centre': job.get('selected_centre'),
#             'created_at': job.get('created_at')
#         } for job_id, job in jobs.items()}
#     })

# @app.route('/api/centres', methods=['GET'])
# def get_centres_alternative():
#     """Alternative endpoint for centres"""
#     return get_centres()

# # ===============================
# # MAIN ENTRY POINT
# # ===============================

# if __name__ == '__main__':
#     # Get port from environment variable (Render sets this)
#     port = int(os.environ.get('PORT', 5000))
    
#     # Check if template files exist
#     if not os.path.exists('CSAT_Treatment_template.xlsx'):
#         print("⚠️ Warning: CSAT_Treatment_template.xlsx not found in current directory")
#     if not os.path.exists('CSAT_Consultation_template.xlsx'):
#         print("⚠️ Warning: CSAT_Consultation_template.xlsx not found in current directory")
    
#     print("="*60)
#     print("🚀 Starting CSAT Reports API Server")
#     print("="*60)
#     print(f"📍 Server running on port: {port}")
#     print(f"📍 Health check: http://localhost:{port}/health")
#     print(f"📍 API endpoints:")
#     print(f"   POST   /api/generate_report - Generate reports")
#     print(f"   GET    /api/job_status/<id> - Check job status")
#     print(f"   GET    /api/get_centres     - Get centre list")
#     print(f"   GET    /api/jobs            - List all jobs")
#     print("="*60)
    
#     app.run(debug=False, host='0.0.0.0', port=port)






# checking email
import numpy as np
import pandas as pd
from flask import Flask, request, jsonify
from flask_cors import CORS
import asyncio
import hashlib
import hmac
import base64
import time
import aiohttp
from aiohttp import ClientTimeout
from urllib.parse import urlparse, parse_qs, urlencode
from datetime import datetime
import os
import re
from openpyxl import load_workbook
from io import BytesIO
import smtplib
from email.message import EmailMessage
import ssl
import requests.utils
import threading
import uuid

app = Flask(__name__)
CORS(app)

# Store job status
jobs = {}

# ===============================
# CONFIG
# ===============================
config = {
    'baseUrl': 'https://3373505-sb1.restlets.api.netsuite.com/app/site/hosting/restlet.nl',
    'realm': '3373505_SB1',
    'consumerKey': 'c6d5e52b1d6b4fa3843170179f35f9eaa725893f03df5191c7b364486bd78253',
    'consumerSecret': 'b3d567c88d518d3db217c6a3c5281684b418a96d471aacd9891d28b4e71394db',
    'token': '82e2861cb21ac4092726b1fe007c736a4616fa72a35e5df4af92e0773541bd32',
    'tokenSecret': '6d36cd7ba2d246100a82c892f543b7eabdea61bcf3a0be691849e42538bcb0e4'
}

# Switched to Port 587 (STARTTLS) which is more reliable on Render/Heroku
EMAIL_CONFIG = {
    'sender': 'channabasava@ayurvaid.com',
    'password': 'tnpyncniqefpcaqn',
    'smtp_server': 'smtp.mail.yahoo.com',
    'smtp_port': 587,
    'receivers': ['channabasavah50@gmail.com']
}

treatment_searches = [
    {'id': 1763, 'name': 'Data-SO'},
    {'id': 1710, 'name': 'Data-TSD+1'},
    {'id': 1712, 'name': 'Data-TED-1'}
]

consultation_searches = [
    {'id': 1719, 'name': 'C+8 Data'},
    {'id': 1764, 'name': 'Data-PP1'}
]

CENTRE_MAPPING_SEARCH_ID = 1765

# ===============================
# HELPER FUNCTIONS
# ===============================
def format_month_year(date_str):
    try:
        date_obj = datetime.strptime(date_str, '%d/%m/%Y')
        return date_obj.strftime('%b-%y')
    except:
        return date_str

def adjust_formula(formula, row):
    return re.sub(r'(\$?[A-Z]+)(\$?)2', lambda m: f"{m.group(1)}{m.group(2)}{row}", formula)

# ===============================
# OAUTH FUNCTIONS
# ===============================
def get_auth_header(url):
    parsed_url = urlparse(url)
    query_params = parse_qs(parsed_url.query)

    oauth_params = {
        'oauth_consumer_key': config['consumerKey'],
        'oauth_nonce': base64.b64encode(os.urandom(32)).decode()[:32],
        'oauth_signature_method': 'HMAC-SHA256',
        'oauth_timestamp': str(int(time.time())),
        'oauth_token': config['token'],
        'oauth_version': '1.0'
    }

    all_params = {k: v[0] for k, v in query_params.items()}
    all_params.update(oauth_params)
    sorted_params = sorted(all_params.items())
    param_string = '&'.join([f"{k}={requests.utils.quote(str(v), safe='')}" for k, v in sorted_params])

    base_string = '&'.join([
        'GET',
        requests.utils.quote(parsed_url.scheme + '://' + parsed_url.netloc + parsed_url.path, safe=''),
        requests.utils.quote(param_string, safe='')
    ])

    signing_key = f"{requests.utils.quote(config['consumerSecret'], safe='')}&{requests.utils.quote(config['tokenSecret'], safe='')}"
    signature = base64.b64encode(hmac.new(signing_key.encode(), base_string.encode(), hashlib.sha256).digest()).decode()
    oauth_params['oauth_signature'] = signature

    auth_header = 'OAuth realm="' + config['realm'] + '", ' + ', '.join([
        f'{k}="{requests.utils.quote(str(v), safe="")}"' for k, v in oauth_params.items()
    ])
    return {'Authorization': auth_header}

# ===============================
# FETCH DATA FUNCTIONS
# ===============================
async def fetch_with_retry(session, url, retries=3):
    for attempt in range(retries):
        try:
            headers = get_auth_header(url)
            headers.update({'Content-Type': 'application/json'})
            async with session.get(url, headers=headers) as res:
                if res.status == 200:
                    return await res.json()
        except Exception as e:
            if attempt == retries - 1: print(f"Fetch failed: {e}")
            await asyncio.sleep(2)
    return None

async def fetch_search(session, search_id, start_date, end_date):
    all_data = []
    page = 1
    while True:
        params = {'script': 559, 'deploy': 1, 'searchId': search_id, 'pageSize': 100, 'pageNumber': page, 'startDate': start_date, 'endDate': end_date}
        url = f"{config['baseUrl']}?{urlencode(params)}"
        res = await fetch_with_retry(session, url)
        if not res or not res.get('results'): break
        results = res['results']
        all_data.extend(results)
        if len(results) < 100: break
        page += 1
    return all_data

async def fetch_centre_mapping(session):
    params = {'script': 559, 'deploy': 1, 'searchId': CENTRE_MAPPING_SEARCH_ID, 'pageSize': 100, 'pageNumber': 1}
    url = f"{config['baseUrl']}?{urlencode(params)}"
    res = await fetch_with_retry(session, url)
    if not res or not res.get('results'): return {}, []
    mapping, names = {}, []
    for item in res['results']:
        c = item.get('AyurVAID Centre', '').strip()
        recs = [item[f'Recipient {i}'].strip() for i in [1, 2] if item.get(f'Recipient {i}')]
        if c and recs:
            mapping[c] = recs
            names.append(c)
    return mapping, sorted(names)

async def fetch_filtered_data(report_type, start_date, end_date):
    timeout = ClientTimeout(total=600) # 10 minute timeout
    async with aiohttp.ClientSession(timeout=timeout) as session:
        centre_mapping, centre_list = await fetch_centre_mapping(session)
        t_data = {s['name']: await fetch_search(session, s['id'], start_date, end_date) for s in treatment_searches} if report_type in ['treatment', 'both'] else {}
        c_data = {s['name']: await fetch_search(session, s['id'], start_date, end_date) for s in consultation_searches} if report_type in ['consultation', 'both'] else {}
        return centre_mapping, centre_list, t_data, c_data

# ===============================
# REPORT & EMAIL FUNCTIONS
# ===============================
def create_report_in_memory(all_data, template_path, sheet_names, start_date=None, target_sheet_for_date=None):
    if not all_data or all(not all_data.get(name) for name in sheet_names): return None
    wb = load_workbook(template_path)
    if start_date and target_sheet_for_date in wb.sheetnames:
        wb[target_sheet_for_date].cell(row=1, column=6, value=format_month_year(start_date))
    
    for name in sheet_names:
        if name not in wb.sheetnames: continue
        ws = wb[name]
        formulas = {c: ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1) if isinstance(ws.cell(row=2, column=c).value, str) and ws.cell(row=2, column=c).value.startswith("=")}
        ws.delete_rows(3, ws.max_row)
        df = pd.DataFrame(all_data.get(name, [])).fillna('')
        for r_idx, row in enumerate(df.values, start=2):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)
            for col, formula in formulas.items():
                ws.cell(row=r_idx, column=col, value=adjust_formula(formula, r_idx))
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def send_email_with_attachment(recipients, subject, body, attachments, job_id=None):
    if not recipients: return False
    msg = EmailMessage()
    msg["Subject"], msg["From"], msg["To"] = subject, EMAIL_CONFIG['sender'], ", ".join(recipients)
    if EMAIL_CONFIG['receivers']: msg["Bcc"] = ", ".join(EMAIL_CONFIG['receivers'])
    msg.set_content(body)
    
    for filename, data in attachments:
        data.seek(0)
        msg.add_attachment(data.read(), maintype="application", subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=filename)

    for attempt in range(3):
        try:
            context = ssl.create_default_context()
            with smtplib.SMTP(EMAIL_CONFIG['smtp_server'], EMAIL_CONFIG['smtp_port'], timeout=30) as smtp:
                smtp.starttls(context=context)
                smtp.login(EMAIL_CONFIG['sender'], EMAIL_CONFIG['password'])
                smtp.send_message(msg)
                return True
        except Exception as e:
            print(f"Email error (Attempt {attempt+1}): {e}")
            time.sleep(5)
    return False

def create_centre_wise_reports(all_data, template_path, sheet_names, report_prefix, timestamp, start_date, target_sheet_for_date=None):
    centre_data = {}
    for sheet_name, data_list in all_data.items():
        if not data_list: continue
        df = pd.DataFrame(data_list)
        if "AyurVAID Centre" not in df.columns: continue
        for centre, group_df in df.groupby("AyurVAID Centre"):
            c_key = str(centre) if not pd.isna(centre) else "Unknown"
            if c_key not in centre_data: centre_data[c_key] = {}
            centre_data[c_key][sheet_name] = group_df.to_dict('records')

    files = []
    for centre, sheets_data in centre_data.items():
        mem_file = create_report_in_memory(sheets_data, template_path, sheet_names, start_date, target_sheet_for_date)
        if mem_file:
            clean_c = str(centre).replace(" ", "_")
            files.append({'centre_name': centre, 'filename': f"{report_prefix}_{clean_c}_{timestamp}.xlsx", 'data': mem_file})
    return files

# ===============================
# BACKGROUND PROCESS
# ===============================
def run_report_generation(report_type, start_date, end_date, selected_centre, job_id):
    try:
        # Create a new event loop for this background thread
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        
        jobs[job_id].update({'status': 'fetching_data', 'progress': 20, 'message': 'Fetching from NetSuite...'})
        mapping, _, t_data, c_data = loop.run_until_complete(fetch_filtered_data(report_type, start_date, end_date))
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        jobs[job_id].update({'status': 'creating_reports', 'progress': 50, 'message': 'Processing Excel files...'})

        t_files = create_centre_wise_reports(t_data, "CSAT_Treatment_template.xlsx", ["Data-SO", "Data-TSD+1", "Data-TED-1"], "Treatment", timestamp, start_date, "CSAT-AVH") if t_data else []
        c_files = create_centre_wise_reports(c_data, "CSAT_Consultation_template.xlsx", ["C+8 Data", "Data-PP1"], "Consultation", timestamp, start_date, "C+8 Dashboard") if c_data else []

        # Sending emails
        jobs[job_id].update({'status': 'sending_emails', 'progress': 70})
        success_count = 0
        
        target_centres = mapping.keys() if selected_centre == 'ALL' else [selected_centre]
        for c_name in target_centres:
            if c_name not in mapping: continue
            atts = []
            for f in t_files: 
                if f['centre_name'] == c_name: atts.append((f['filename'], f['data']))
            for f in c_files: 
                if f['centre_name'] == c_name: atts.append((f['filename'], f['data']))
            
            if atts:
                subj = f"CSAT Reports - {c_name} ({start_date}-{end_date})"
                body = f"Attached are reports for {c_name}."
                if send_email_with_attachment(mapping[c_name], subj, body, atts): success_count += 1
        
        jobs[job_id].update({'status': 'completed', 'progress': 100, 'message': f'Finished. Emails sent to {success_count} centres.'})
    except Exception as e:
        jobs[job_id].update({'status': 'failed', 'message': str(e)})

# ===============================
# API ROUTES
# ===============================
@app.route('/api/generate_report', methods=['POST'])
def generate_report():
    data = request.json
    job_id = str(uuid.uuid4())
    jobs[job_id] = {'status': 'pending', 'created_at': datetime.now().isoformat(), 'report_type': data.get('report_type')}
    
    t = threading.Thread(target=run_report_generation, args=(data['report_type'], data['start_date'], data['end_date'], data.get('selected_centre', 'ALL'), job_id))
    t.daemon = True
    t.start()
    return jsonify({'job_id': job_id, 'status': 'started'})

@app.route('/api/job_status/<job_id>')
def job_status(job_id):
    return jsonify(jobs.get(job_id, {'error': 'Job not found'}))

@app.route('/health')
def health():
    return jsonify({'status': 'ok'})

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
